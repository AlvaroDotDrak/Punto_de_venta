import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models import Expense, ExpenseCategory, Invoice, Sale, SaleItem, Seller, Product, IngredientMovement, ProductRecipe
from ..auth import require_admin
from ..audit import ACTIONS, log_action
from ..schemas import AccountingSummary, ExpenseSummaryItem, IncomeSummaryItem, LossesReport, LossSummaryItem, LossReasonItem

from ..utils import calculate_vat, compute_cost_per_unit

router = APIRouter(prefix="/accounting", tags=["accounting"])


MONTH_NAMES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# Nombres para mostrar en reportes
_CAT_LABELS = {
    "vitrina": "Vitrina",
    "salados": "Salados",
    "bebidas": "Bebidas",
    "cafe": "Café",
    "encargo": "Encargos",
    "mostrador": "Mostrador",
}


@router.get("/summary", response_model=AccountingSummary)
def get_summary(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to + "T23:59:59")

    # ── Ventas ───────────────────────────────────────────────────────────────
    sales = db.query(Sale).filter(
        Sale.status == "completed",
        Sale.created_at >= dt_from,
        Sale.created_at <= dt_to,
    ).all()

    total_income = sum(s.total for s in sales)
    total_income_cash = sum(s.total for s in sales if s.payment_method == "efectivo")
    total_income_card = sum(s.total for s in sales if s.payment_method == "tarjeta")
    total_income_transfer = sum(s.total for s in sales if s.payment_method == "transferencia")
    sales_count = len(sales)
    sales_with_receipt = sum(1 for s in sales if s.has_receipt)

    # ── Gastos ───────────────────────────────────────────────────────────────
    expenses = db.query(Expense).options(
        joinedload(Expense.purchase_items)
    ).filter(
        Expense.created_at >= dt_from,
        Expense.created_at <= dt_to,
    ).all()
    total_expenses = sum(e.amount for e in expenses)

    cat_map: dict[str, dict] = {}
    cat_names: dict[int, str] = {}

    def _cat_name(cid: int) -> str:
        if cid not in cat_names:
            cat = db.query(ExpenseCategory).filter(ExpenseCategory.id == cid).first()
            cat_names[cid] = cat.name if cat else "Sin categoría"
        return cat_names[cid]

    # Una compra con líneas en distintas categorías reparte su monto (IVA incluido)
    # entre cada categoría según el peso neto de sus líneas. Sin líneas → cabecera.
    for e in expenses:
        items = e.purchase_items
        net = sum(it.line_total for it in items) if items else 0.0
        if items and net > 0:
            shares: dict[int, float] = {}
            for it in items:
                cid = it.category_id or e.category_id
                shares[cid] = shares.get(cid, 0.0) + it.line_total
            for cid, line_net in shares.items():
                slot = cat_map.setdefault(_cat_name(cid), {"total": 0.0, "ids": set()})
                slot["total"] += e.amount * (line_net / net)
                slot["ids"].add(e.id)
        else:
            slot = cat_map.setdefault(_cat_name(e.category_id), {"total": 0.0, "ids": set()})
            slot["total"] += e.amount
            slot["ids"].add(e.id)

    expenses_by_category = [
        ExpenseSummaryItem(category_name=name, total=round(v["total"]), count=len(v["ids"]))
        for name, v in sorted(cat_map.items(), key=lambda x: x[1]["total"], reverse=True)
    ]

    # ── Facturas emitidas ─────────────────────────────────────────────────────
    invoices = db.query(Invoice).filter(
        Invoice.created_at >= dt_from,
        Invoice.created_at <= dt_to,
    ).all()
    invoices_tax_total = sum(i.tax_amount for i in invoices)

    # ── IVA estimado (extracción 19/119 — precios incluyen IVA) ──────────────
    # Débito Fiscal: IVA cobrado en ventas con boleta + IVA de facturas emitidas
    sales_with_receipt_total = sum(s.total for s in sales if s.has_receipt)
    vat_debit = calculate_vat(sales_with_receipt_total) + invoices_tax_total

    # Crédito Fiscal: IVA en compras donde el proveedor emitió factura
    expenses_factura_total = sum(e.amount for e in expenses if (e.document_type or 'boleta') == 'factura')
    vat_credit = calculate_vat(expenses_factura_total)

    vat_balance = vat_debit - vat_credit

    # ── Ingresos por categoría de producto ───────────────────────────────────
    sale_items = (
        db.query(SaleItem)
        .join(Sale, SaleItem.sale_id == Sale.id)
        .filter(
            Sale.status == "completed",
            Sale.created_at >= dt_from,
            Sale.created_at <= dt_to,
        )
        .options(joinedload(SaleItem.product))
        .all()
    )

    income_cat: dict[str, dict] = {}
    for item in sale_items:
        raw_cat = item.product.category if item.product else "Sin categoría"
        label = _CAT_LABELS.get(raw_cat, raw_cat.capitalize())
        if label not in income_cat:
            income_cat[label] = {"total": 0.0, "count": 0}
        income_cat[label]["total"] += item.subtotal
        income_cat[label]["count"] += item.quantity

    income_by_category = [
        IncomeSummaryItem(category_name=name, total=v["total"], count=v["count"])
        for name, v in sorted(income_cat.items(), key=lambda x: x[1]["total"], reverse=True)
    ]

    return AccountingSummary(
        date_from=date_from,
        date_to=date_to,
        total_income=total_income,
        total_income_cash=total_income_cash,
        total_income_card=total_income_card,
        total_income_transfer=total_income_transfer,
        total_expenses=total_expenses,
        net_profit=total_income - total_expenses,
        sales_count=sales_count,
        sales_with_receipt=sales_with_receipt,
        sales_without_receipt=sales_count - sales_with_receipt,
        expenses_by_category=expenses_by_category,
        invoices_count=len(invoices),
        invoices_total=sum(i.total_amount for i in invoices),
        vat_debit=vat_debit,
        vat_credit=vat_credit,
        vat_balance=vat_balance,
        income_by_category=income_by_category,
    )


@router.get("/export")
def export_report(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to + "T23:59:59")

    sales = db.query(Sale).filter(
        Sale.status == "completed",
        Sale.created_at >= dt_from,
        Sale.created_at <= dt_to,
    ).all()
    expenses = db.query(Expense).filter(
        Expense.created_at >= dt_from,
        Expense.created_at <= dt_to,
    ).all()
    invoices = db.query(Invoice).filter(
        Invoice.created_at >= dt_from,
        Invoice.created_at <= dt_to,
    ).all()

    total_income = sum(s.total for s in sales)
    total_expenses = sum(e.amount for e in expenses)
    period_label = f"{date_from} al {date_to}"

    sales_with_receipt_total = sum(s.total for s in sales if s.has_receipt)
    invoices_tax_total = sum(i.tax_amount for i in invoices)
    vat_debit = calculate_vat(sales_with_receipt_total) + invoices_tax_total
    vat_credit = calculate_vat(sum(e.amount for e in expenses if (e.document_type or 'boleta') == 'factura'))
    vat_balance = vat_debit - vat_credit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="BF5A2F")

    def style_header_row(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append(["Período", "Total Ingresos", "Total Gastos", "Utilidad Neta",
                "Ventas c/boleta", "Ventas s/boleta",
                "Débito Fiscal IVA", "Crédito Fiscal IVA", "IVA Estimado a Pagar"])
    style_header_row(ws1)
    ws1.append([
        period_label,
        total_income,
        total_expenses,
        total_income - total_expenses,
        sum(1 for s in sales if s.has_receipt),
        sum(1 for s in sales if not s.has_receipt),
        round(vat_debit),
        round(vat_credit),
        round(vat_balance),
    ])
    for col, width in [("A", 22), ("B", 16), ("C", 14), ("D", 14),
                       ("E", 15), ("F", 15), ("G", 18), ("H", 18), ("I", 20)]:
        ws1.column_dimensions[col].width = width

    # ── Hoja 2: Detalle Ventas ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle Ventas")
    ws2.append(["Fecha", "Monto", "Método Pago", "Con Boleta", "Vendedor", "Estado"])
    style_header_row(ws2)

    seller_cache: dict[int, str] = {}
    for s in sorted(sales, key=lambda x: x.created_at):
        if s.seller_id not in seller_cache:
            sel = db.query(Seller).filter(Seller.id == s.seller_id).first()
            seller_cache[s.seller_id] = sel.name if sel else "Desconocido"
        ws2.append([
            s.created_at.strftime("%d/%m/%Y %H:%M"),
            s.total,
            s.payment_method,
            "Sí" if s.has_receipt else "No",
            seller_cache[s.seller_id],
            s.status,
        ])
    for col, width in [("A", 18), ("B", 12), ("C", 16), ("D", 12), ("E", 18), ("F", 12)]:
        ws2.column_dimensions[col].width = width

    # ── Hoja 3: Detalle Gastos ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Detalle Gastos")
    ws3.append(["Fecha", "Monto", "Tipo Doc.", "Categoría", "Descripción", "Registrado por"])
    style_header_row(ws3)

    cat_cache: dict[int, str] = {}
    for e in sorted(expenses, key=lambda x: x.created_at):
        if e.category_id not in cat_cache:
            cat = db.query(ExpenseCategory).filter(ExpenseCategory.id == e.category_id).first()
            cat_cache[e.category_id] = cat.name if cat else "Sin categoría"
        if e.seller_id not in seller_cache:
            sel = db.query(Seller).filter(Seller.id == e.seller_id).first()
            seller_cache[e.seller_id] = sel.name if sel else "Desconocido"
        ws3.append([
            e.created_at.strftime("%d/%m/%Y %H:%M"),
            e.amount,
            (e.document_type or 'boleta').capitalize(),
            cat_cache[e.category_id],
            e.description or "",
            seller_cache[e.seller_id],
        ])
    for col, width in [("A", 18), ("B", 12), ("C", 12), ("D", 18), ("E", 30), ("F", 18)]:
        ws3.column_dimensions[col].width = width

    # ── Hoja 4: Detalle Mermas ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Detalle Mermas")
    ws4.append(["Fecha", "Insumo", "Cantidad", "Unidad", "Costo Pérdida", "Motivo / Notas", "Registrado por"])
    style_header_row(ws4)

    mermas = db.query(IngredientMovement).options(
        joinedload(IngredientMovement.ingredient)
    ).filter(
        IngredientMovement.type == "loss",
        IngredientMovement.created_at >= dt_from,
        IngredientMovement.created_at <= dt_to,
    ).all()

    for m in sorted(mermas, key=lambda x: x.created_at):
        if m.seller_id and m.seller_id not in seller_cache:
            sel = db.query(Seller).filter(Seller.id == m.seller_id).first()
            seller_cache[m.seller_id] = sel.name if sel else "Desconocido"
        
        cost_loss = m.cost if m.cost is not None else (m.quantity * (m.ingredient.last_price if m.ingredient else 0.0))
        ing_name = m.ingredient.name if m.ingredient else f"Insumo #{m.ingredient_id}"
        ing_unit = m.ingredient.unit if m.ingredient else ""

        ws4.append([
            m.created_at.strftime("%d/%m/%Y %H:%M"),
            ing_name,
            m.quantity,
            ing_unit,
            round(cost_loss, 2),
            m.notes or "",
            seller_cache.get(m.seller_id, "Desconocido") if m.seller_id else "Desconocido",
        ])
    for col, width in [("A", 18), ("B", 20), ("C", 12), ("D", 12), ("E", 14), ("F", 30), ("G", 18)]:
        ws4.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    log_action(db, ACTIONS.ACCOUNTING_EXPORT, admin.id, f"Exportación contable {period_label}")

    filename = f"reporte_contable_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/losses", response_model=LossesReport)
def get_losses_report(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to + "T23:59:59")

    # Obtener todos los movimientos de tipo "loss"
    movements = (
        db.query(IngredientMovement)
        .options(joinedload(IngredientMovement.ingredient))
        .filter(
            IngredientMovement.type == "loss",
            IngredientMovement.created_at >= dt_from,
            IngredientMovement.created_at <= dt_to,
        )
        .all()
    )

    # Agrupar por ingrediente
    by_ing = {}
    # Agrupar por motivo (notes)
    by_reason = {}
    
    total_loss_cost = 0.0

    for m in movements:
        # Si no tiene costo histórico, usar el fallback del last_price
        cost = m.cost if m.cost is not None else (m.quantity * (m.ingredient.last_price if m.ingredient else 0.0))
        total_loss_cost += cost
        
        # Agrupar ingrediente
        ing_id = m.ingredient_id
        ing_name = m.ingredient.name if m.ingredient else f"Insumo #{ing_id}"
        ing_unit = m.ingredient.unit if m.ingredient else ""
        if ing_id not in by_ing:
            by_ing[ing_id] = {
                "ingredient_id": ing_id,
                "name": ing_name,
                "quantity": 0.0,
                "unit": ing_unit,
                "total_cost": 0.0
            }
        by_ing[ing_id]["quantity"] += m.quantity
        by_ing[ing_id]["total_cost"] += cost

        # Agrupar motivo (normalizar a "Sin descripción" si es vacío)
        reason = m.notes.strip() if m.notes else "Sin descripción"
        if reason not in by_reason:
            by_reason[reason] = {
                "notes": reason,
                "total_cost": 0.0,
                "count": 0
            }
        by_reason[reason]["total_cost"] += cost
        by_reason[reason]["count"] += 1

    # Convertir dicts a listas y ordenar por costo descendente
    by_ingredient_list = [
        LossSummaryItem(
            ingredient_id=item["ingredient_id"],
            name=item["name"],
            quantity=round(item["quantity"], 2),
            unit=item["unit"],
            total_cost=round(item["total_cost"], 2)
        )
        for item in by_ing.values()
    ]
    by_ingredient_list.sort(key=lambda x: x.total_cost, reverse=True)

    by_reason_list = [
        LossReasonItem(
            notes=item["notes"],
            total_cost=round(item["total_cost"], 2),
            count=item["count"]
        )
        for item in by_reason.values()
    ]
    by_reason_list.sort(key=lambda x: x.total_cost, reverse=True)

    return LossesReport(
        date_from=date_from,
        date_to=date_to,
        total_loss_cost=round(total_loss_cost, 2),
        by_ingredient=by_ingredient_list,
        by_reason=by_reason_list
    )


@router.get("/profitability")
def get_profitability(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: Session = Depends(get_db),
    admin=Depends(require_admin),
):
    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to + "T23:59:59")

    sale_items = (
        db.query(SaleItem)
        .join(Sale)
        .filter(
            Sale.status == "completed",
            Sale.created_at >= dt_from,
            Sale.created_at <= dt_to
        )
        .all()
    )

    # Agrupar por (product_id, showcase_type) para separar enteros de trozos
    grouped = {}
    for item in sale_items:
        if not item.product_id:
            continue
        key = (item.product_id, item.showcase_type)
        if key not in grouped:
            suffix = " (entero)" if item.showcase_type == "entero" else " (trozo)" if item.showcase_type == "trozado" else ""
            grouped[key] = {
                "product_id": item.product_id,
                "name": item.product_name + suffix,
                "showcase_type": item.showcase_type,
                "units_sold": 0,
                "revenue": 0.0,
            }
        grouped[key]["units_sold"] += item.quantity
        grouped[key]["revenue"] += item.subtotal

    product_results = []
    total_revenue = 0.0
    total_cogs = 0.0

    if grouped:
        product_ids = list({k[0] for k in grouped.keys()})
        products = db.query(Product).options(
            joinedload(Product.recipes).joinedload(ProductRecipe.ingredient)
        ).filter(Product.id.in_(product_ids)).all()

        product_map = {p.id: p for p in products}

        for (p_id, showcase_type), data in grouped.items():
            product = product_map.get(p_id)
            if not product:
                continue

            cost_per_unit = compute_cost_per_unit(product, showcase_type)

            if cost_per_unit is None:
                continue  # excluir productos sin costo definido

            data["category"] = product.category
            data["has_recipe"] = len(product.recipes) > 0

            cogs = round(cost_per_unit * data["units_sold"], 2)
            profit = round(data["revenue"] - cogs, 2)
            margin = round((profit / data["revenue"]) * 100, 1) if data["revenue"] > 0 else 0.0

            data["cogs"] = cogs
            data["profit"] = profit
            data["margin"] = margin

            total_revenue += data["revenue"]
            total_cogs += cogs

            product_results.append(data)

    product_results.sort(key=lambda x: x["profit"], reverse=True)

    cat_map = {}
    for p in product_results:
        c = p["category"]
        if c not in cat_map:
            cat_map[c] = {"label": _CAT_LABELS.get(c, c.capitalize()), "revenue": 0.0, "cogs": 0.0}
        cat_map[c]["revenue"] += p["revenue"]
        cat_map[c]["cogs"] += p["cogs"]

    categories = []
    for c, stats in cat_map.items():
        c_prof = stats["revenue"] - stats["cogs"]
        stats["margin"] = round((c_prof / stats["revenue"]) * 100, 1) if stats["revenue"] > 0 else 0
        categories.append(stats)

    categories.sort(key=lambda x: x["revenue"], reverse=True)

    total_profit = total_revenue - total_cogs
    total_margin = round((total_profit / total_revenue) * 100, 1) if total_revenue > 0 else 0

    return {
        "total_revenue": round(total_revenue, 2),
        "total_cogs": round(total_cogs, 2),
        "total_profit": round(total_profit, 2),
        "total_margin": total_margin,
        "categories": categories,
        "products": product_results,
    }

